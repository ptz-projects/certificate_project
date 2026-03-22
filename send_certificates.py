"""
send_certificates.py
--------------------
Stage 2: Email generated PDF certificates to attendees via Microsoft Graph API.
Uses OAuth2 client credentials flow — works with Security Defaults enabled.
Run ONLY after generate_certificates.py has completed successfully.

Requirements:
    pip install msal requests openpyxl python-dotenv

Credentials are loaded from .env file — never hardcoded here.
"""

import os
import base64
import openpyxl
import requests
import msal
from dotenv import load_dotenv

# ─────────────────────────────────────────────
# LOAD CREDENTIALS FROM .env
# ─────────────────────────────────────────────
load_dotenv()

TENANT_ID = os.getenv("AZURE_TENANT_ID")
CLIENT_ID = os.getenv("AZURE_CLIENT_ID")
CLIENT_SECRET = os.getenv("AZURE_CLIENT_SECRET")
SENDER_EMAIL = os.getenv("SENDER_EMAIL")
BUSINESS_NAME = os.getenv("BUSINESS_NAME")

# ─────────────────────────────────────────────
# CONFIGURATION
# ─────────────────────────────────────────────
CERTIFICATES_FOLDER = "certificates"
EXCEL_FILE = "attendees.xlsx"
GRAPH_SCOPE = ["https://graph.microsoft.com/.default"]
GRAPH_ENDPOINT = f"https://graph.microsoft.com/v1.0/users/{SENDER_EMAIL}/sendMail"

EMAIL_SUBJECT = "Your Certificate of Completion"
EMAIL_BODY_TEMPLATE = """
Dear {first_name},

Thank you for attending {course_name}.

Please find attached your Certificate of Completion.

We hope you found the session valuable and look forward to seeing you at future events.

Warm regards,
{business_name}
""".strip()


# ─────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────

def validate_config():
    """Ensure all required environment variables are loaded."""
    missing = []
    for name, value in [
        ("AZURE_TENANT_ID", TENANT_ID),
        ("AZURE_CLIENT_ID", CLIENT_ID),
        ("AZURE_CLIENT_SECRET", CLIENT_SECRET),
        ("SENDER_EMAIL", SENDER_EMAIL),
        ("BUSINESS_NAME", BUSINESS_NAME),
    ]:
        if not value or value.startswith("your-"):
            missing.append(name)
    if missing:
        print("\n✗ Missing or unconfigured .env values:")
        for m in missing:
            print(f"  - {m}")
        print("\nEdit your .env file with real values before running.\n")
        return False
    return True


def get_access_token():
    """Obtain OAuth2 access token from Microsoft identity platform."""
    authority = f"https://login.microsoftonline.com/{TENANT_ID}"
    app = msal.ConfidentialClientApplication(
        CLIENT_ID,
        authority=authority,
        client_credential=CLIENT_SECRET
    )
    result = app.acquire_token_for_client(scopes=GRAPH_SCOPE)
    if "access_token" in result:
        print("  ✓ OAuth2 token acquired")
        return result["access_token"]
    else:
        error = result.get("error_description", "Unknown error")
        raise Exception(f"Token acquisition failed: {error}")


def load_attendees(filepath):
    """Load attendee data from Excel file."""
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    attendees = []
    headers = [cell.value for cell in ws[1]]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(row):
            attendee = dict(zip(headers, row))
            attendees.append(attendee)
    return attendees


def find_certificate(name, folder):
    """Find the generated PDF for a given attendee name."""
    safe_name = name.replace(" ", "_").replace(".", "_")
    filename = f"{safe_name}_certificate.pdf"
    filepath = os.path.join(folder, filename)
    return filepath if os.path.exists(filepath) else None


def build_email_payload(attendee, cert_path):
    """Build the Graph API email payload with PDF attachment."""
    full_name = attendee.get("first.lastname", "")
    course_name = attendee.get("webinar course name", "")
    email = attendee.get("email", "")

    # First name for greeting
    first_name = full_name.split(".")[0] if "." in full_name else full_name.split(" ")[0]

    # Encode PDF as base64
    with open(cert_path, "rb") as f:
        pdf_base64 = base64.b64encode(f.read()).decode("utf-8")

    cert_filename = os.path.basename(cert_path)

    body_text = EMAIL_BODY_TEMPLATE.format(
        first_name=first_name,
        course_name=course_name,
        business_name=BUSINESS_NAME
    )

    return {
        "message": {
            "subject": EMAIL_SUBJECT,
            "body": {
                "contentType": "Text",
                "content": body_text
            },
            "toRecipients": [
                {"emailAddress": {"address": email}}
            ],
            "attachments": [
                {
                    "@odata.type": "#microsoft.graph.fileAttachment",
                    "name": cert_filename,
                    "contentType": "application/pdf",
                    "contentBytes": pdf_base64
                }
            ]
        },
        "saveToSentItems": "true"
    }


def send_email(token, attendee, cert_path):
    """Send a single certificate email via Microsoft Graph API."""
    payload = build_email_payload(attendee, cert_path)
    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json"
    }
    response = requests.post(GRAPH_ENDPOINT, headers=headers, json=payload)

    if response.status_code == 202:
        name = attendee.get("first.lastname", "")
        email = attendee.get("email", "")
        print(f"  ✓ Sent to: {email} ({name})")
        return True
    else:
        name = attendee.get("first.lastname", "")
        print(f"  ✗ Failed for {name}: {response.status_code} — {response.text}")
        return False


# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────

def main():
    print("\n── Certificate Sender (Microsoft Graph API) ──")

    # Validate config
    if not validate_config():
        return

    print(f"Sender:  {SENDER_EMAIL}")
    print(f"Business: {BUSINESS_NAME}\n")

    # Load attendees
    attendees = load_attendees(EXCEL_FILE)
    print(f"Found {len(attendees)} attendees")

    # Verify all certificates exist before attempting to send
    print("Verifying certificates...")
    ready = []
    missing = []
    for attendee in attendees:
        name = attendee.get("first.lastname", "")
        cert = find_certificate(name, CERTIFICATES_FOLDER)
        if cert:
            ready.append((attendee, cert))
        else:
            missing.append(name)

    if missing:
        print(f"\n  ✗ Missing certificates for:")
        for name in missing:
            print(f"    - {name}")
        print("\nRun generate_certificates.py first. Aborting.\n")
        return

    print(f"All {len(ready)} certificates found.\n")

    # Acquire OAuth2 token
    print("Authenticating with Microsoft Graph...")
    try:
        token = get_access_token()
    except Exception as e:
        print(f"\n✗ Authentication error: {e}")
        print("\nCheck your AZURE_TENANT_ID, AZURE_CLIENT_ID, and AZURE_CLIENT_SECRET in .env\n")
        return

    # Send emails
    print("\nSending certificates...\n")
    sent = 0
    failed = []

    for attendee, cert_path in ready:
        success = send_email(token, attendee, cert_path)
        if success:
            sent += 1
        else:
            failed.append(attendee.get("first.lastname", "unknown"))

    # Summary
    print(f"\n── Complete ──")
    print(f"  Sent:   {sent}")
    print(f"  Failed: {len(failed)}")
    if failed:
        print("  Failed recipients:")
        for name in failed:
            print(f"    - {name}")
    print()


if __name__ == "__main__":
    main()
