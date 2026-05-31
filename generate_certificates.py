# Stage 1 — Generate PDF certificates
"""
generate_certificates.py
------------------------
Stage 1: Generate branded PDF certificates from Excel attendee list.
Outputs one PDF per attendee into the /certificates folder.

Requirements:
    pip install reportlab openpyxl pillow requests

Project structure expected:
    /certificate_project
        generate_certificates.py
        send_certificates.py
        attendees.xlsx
        assets/
            logo.png          <-- Add your logo here
        certificates/         <-- Generated PDFs will appear here
"""

import os
import openpyxl
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas
from reportlab.lib.utils import ImageReader
import urllib.request
import io
from dotenv import load_dotenv

load_dotenv()

# ─────────────────────────────────────────────
# CONFIGURATION — Edit these values
# ─────────────────────────────────────────────
BUSINESS_NAME = os.getenv("BUSINESS_NAME")       # Issued by: name on certificate
BRAND_COLOUR = colors.HexColor("#1A3C5E")     # Replace with your brand hex colour
ACCENT_COLOUR = colors.HexColor("#C9A84C")    # Replace with your accent/gold colour
LOGO_PATH = "assets/logo.png"                 # Path to your logo file
OUTPUT_FOLDER = "certificates"                # Where PDFs are saved
EXCEL_FILE = "attendees.xlsx"                 # Your attendee list

# Medal image — uses a royalty-free placeholder emoji-style medal via local drawing
# Replace with your own medal image by setting MEDAL_PATH = "assets/medal.png"
MEDAL_PATH = None  # Set to "assets/medal.png" if you have one

# ─────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────

def draw_medal(c, x, y, size=60):
    """Draws a simple decorative medal if no medal image is provided."""
    # Outer circle
    c.setFillColor(ACCENT_COLOUR)
    c.circle(x, y, size / 2 + 8, fill=1, stroke=0)
    # Inner circle
    c.setFillColor(colors.HexColor("#FFD700"))
    c.circle(x, y, size / 2, fill=1, stroke=0)
    # Star shape (simplified)
    c.setFillColor(ACCENT_COLOUR)
    c.setFont("Helvetica-Bold", 28)
    c.drawCentredString(x, y - 10, "★")


def load_attendees(filepath):
    """Load attendee data from Excel file."""
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    attendees = []
    headers = [cell.value for cell in ws[1]]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(row):  # Skip empty rows
            attendee = dict(zip(headers, row))
            attendees.append(attendee)
    return attendees


def generate_certificate(attendee, output_folder):
    """Generate a single branded PDF certificate for one attendee."""
    
    # Extract fields — matches your Excel columns
    email = attendee.get("email", "")
    full_name = attendee.get("full name", "")
    course_name = attendee.get("course name", "")
    completion_date = attendee.get("date", "")
    
    # Format date if it's a datetime object
    if hasattr(completion_date, 'strftime'):
        completion_date = completion_date.strftime("%d %B %Y")
    else:
        completion_date = str(completion_date)

    # Safe filename from name
    safe_name = full_name.replace(" ", "_").replace(".", "_")
    filename = f"{safe_name}_certificate.pdf"
    filepath = os.path.join(output_folder, filename)

    # Page setup — landscape A4
    width, height = landscape(A4)
    c = canvas.Canvas(filepath, pagesize=landscape(A4))

    # ── Background ──
    c.setFillColor(colors.white)
    c.rect(0, 0, width, height, fill=1, stroke=0)

    # ── Decorative border ──
    c.setStrokeColor(BRAND_COLOUR)
    c.setLineWidth(3)
    c.rect(15*mm, 15*mm, width - 30*mm, height - 30*mm, fill=0, stroke=1)
    c.setStrokeColor(ACCENT_COLOUR)
    c.setLineWidth(1)
    c.rect(18*mm, 18*mm, width - 36*mm, height - 36*mm, fill=0, stroke=1)

    # ── Logo top left ──
    if os.path.exists(LOGO_PATH):
        logo = ImageReader(LOGO_PATH)
        c.drawImage(logo, 25*mm, height - 45*mm, width=45*mm, height=20*mm,
                    preserveAspectRatio=True, mask='auto')

    # ── Certificate of Completion heading ──
    c.setFillColor(BRAND_COLOUR)
    c.setFont("Helvetica-Bold", 32)
    c.drawCentredString(width / 2, height - 62*mm, "Certificate of Participation")

    # ── Decorative line under heading ──
    c.setStrokeColor(ACCENT_COLOUR)
    c.setLineWidth(2)
    c.line(width/2 - 80*mm, height - 67*mm, width/2 + 80*mm, height - 67*mm)

    # ── Medal centre ──
    medal_y = height / 2 + 5*mm
    if MEDAL_PATH and os.path.exists(MEDAL_PATH):
        medal_img = ImageReader(MEDAL_PATH)
        c.drawImage(medal_img, width/2 - 20*mm, medal_y - 20*mm,
                    width=40*mm, height=40*mm, preserveAspectRatio=True, mask='auto')
    else:
        draw_medal(c, width / 2, medal_y + 10*mm, size=55)

    # ── This certifies that ──
    c.setFillColor(colors.HexColor("#555555"))
    c.setFont("Helvetica", 14)
    c.drawCentredString(width / 2, height / 2 - 18*mm, "This certifies that")

    # ── Attendee name ──
    c.setFillColor(BRAND_COLOUR)
    c.setFont("Helvetica-Bold", 26)
    c.drawCentredString(width / 2, height / 2 - 30*mm, full_name)

    # ── Name underline ──
    c.setStrokeColor(ACCENT_COLOUR)
    c.setLineWidth(1)
    name_width = c.stringWidth(full_name, "Helvetica-Bold", 26)
    c.line(width/2 - name_width/2, height/2 - 32*mm,
           width/2 + name_width/2, height/2 - 32*mm)

    # ── Has completed the course ──
    c.setFillColor(colors.HexColor("#555555"))
    c.setFont("Helvetica", 13)
    c.drawCentredString(width / 2, height / 2 - 42*mm, "participated in the course")


    # ── Course name ──
    c.setFillColor(BRAND_COLOUR)
    c.setFont("Helvetica-Bold", 16)
    c.drawCentredString(width / 2, height / 2 - 53*mm, course_name)

    # ── Bottom row: completion date left, issued by right ──
    bottom_y = 28*mm
    c.setFillColor(colors.HexColor("#555555"))
    c.setFont("Helvetica", 11)
    c.drawString(25*mm, bottom_y + 5*mm, "Completion Date")
    c.setFont("Helvetica-Bold", 12)
    c.setFillColor(BRAND_COLOUR)
    c.drawString(25*mm, bottom_y - 3*mm, completion_date)

    c.setFillColor(colors.HexColor("#555555"))
    c.setFont("Helvetica", 11)
    c.drawRightString(width - 25*mm, bottom_y + 5*mm, "Issued by:")
    c.setFont("Helvetica-Bold", 12)
    c.setFillColor(BRAND_COLOUR)
    c.drawRightString(width - 25*mm, bottom_y - 3*mm, BUSINESS_NAME)

    c.save()
    print(f"  ✓ Generated: {filename}")
    return filepath, email, full_name


def main():
    print("\n── Certificate Generator ──")
    print(f"Reading attendees from: {EXCEL_FILE}")

    # Create output folder
    os.makedirs(OUTPUT_FOLDER, exist_ok=True)

    # Load attendees
    attendees = load_attendees(EXCEL_FILE)
    print(f"Found {len(attendees)} attendees\n")

    # Generate each certificate
    results = []
    for attendee in attendees:
        try:
            filepath, email, name = generate_certificate(attendee, OUTPUT_FOLDER)
            results.append({"name": name, "email": email, "file": filepath})
        except Exception as e:
            print(f"  ✗ Error for {attendee}: {e}")

    print(f"\n── Complete: {len(results)} certificates saved to /{OUTPUT_FOLDER} ──")
    print("Run send_certificates.py when ready to email.\n")


if __name__ == "__main__":
    main()
